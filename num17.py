# -*- coding: utf-8 -*-
"""
@Auth ： xiaobaihe
@File ：num1.py
@Time ：2021/6/9 
"""

import openpyxl
import requests

def func(url, data, headers={'X-Lemonban-Media-Type': 'lemonban.v2','Content-Type': 'application/json'}):
    res = requests.post(url=url, json=data, headers=headers)
    res_res = res.json()
    return res_res                                   #调用函数发送请求

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook('test_case_api.xlsx')
    sheet = wb['register']
    row = sheet.max_row
    list_1 = []
    for item in range(2,row+1):
        dict_1 = dict(
        id_reg = sheet.cell(row=item,column=1).value,
        url_reg = sheet.cell(row=item,column=5).value,
        data_reg = sheet.cell(row=item,column=6).value,
        expected_reg = sheet.cell(row=item,column=7).value)
        list_1.append(dict_1)
    return list_1


def write_data(filename, sheetname, row, final_result, column=8):  # 列设为默认值
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)
    wb.close()



res =read_data('test_case_api.xlsx','register')   #调用函数读取 List 的测试用例
print(res)

for case in res:
    # print(case)
    case_id = case['id_reg']                    #取出用例编号
    case_url = case.get('url_reg')              #取 url
    case_data = case.get('data_reg')            #取请求参数
    case_expect = case['expected_reg']          #取出预期结果
    case_data = eval(case_data)       #通过eval函数，将取出的字符串格式的data，转换成字典格式的data
    case_expect = eval(case_expect)            #转换预期结果
    real_result = func(url= case_url,data = case_data)
    # print(real_result)
    case_expect_msg = case_expect['msg']      #预期结果的 ‘msg'
    real_result_msg = real_result['msg']      #实际结果的 ‘msg"
    print(case_expect_msg,real_result_msg)
    print('用例编号:{}'.format(case_id))
    print('预期结果为:{}'.format(case_expect_msg))
    print('实际结果为:{}'.format(real_result_msg))
    if case_expect_msg == real_result_msg:
        print('这条用例通过!!')
        final_result = 'pass'        #设置变量来接收最终的结果传给写入函数
    else:
        print('这条用例不通过!!')
        final_result = 'false'
    print('*'*50)
    write_data('test_case_api.xlsx','register',case_id+1,final_result,8)



    print(case_id,case_url,case_data,case_expect)


